from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image
from tests.runtime_support import configure_test_runtime


@pytest.fixture()
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    configure_test_runtime(tmp_path, database_name="client-platform.db")
    from app.main import app

    with TestClient(app) as test_client:
        admin = test_client.post(
            "/api/auth/setup",
            json={
                "brand_name": "测试商标",
                "username": "admin",
                "display_name": "测试管理员",
                "password": "AdminPass123",
            },
        )
        assert admin.status_code == 201, admin.text
        admin_data = admin.json()["data"]
        test_client.admin_password = admin_data["generated_password"]
        test_client.headers["X-CSRF-Token"] = admin_data["csrf_token"]
        yield test_client


def unwrap(response, status: int | None = None):
    if status is not None:
        assert response.status_code == status, response.text
    else:
        assert response.status_code < 400, response.text
    body = response.json()
    assert body["success"] is True, body
    return body["data"]


def register(client: TestClient, suffix: str):
    response = client.post(
        "/api/client/auth/register",
        json={
            "username": f"client_{suffix}",
            "phone": f"1380000{int(suffix):04d}",
            "nickname": f"客户{suffix}",
            "email": f"client{suffix}@example.com",
            "password": "ClientPass123",
        },
    )
    data = unwrap(response, 201)
    client.headers["X-CSRF-Token"] = data["csrf_token"]
    return data, client.cookies.get("service_client_session")


def admin_headers(client: TestClient) -> dict[str, str]:
    login = client.post(
        "/api/auth/login", json={"username": "admin", "password": client.admin_password}
    )
    data = unwrap(login)
    return {"X-CSRF-Token": data["csrf_token"]}


def png_bytes() -> bytes:
    stream = BytesIO()
    Image.new("RGB", (16, 16), "#3366ff").save(stream, format="PNG")
    return stream.getvalue()


def recycle_catalog_xlsx(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["品牌", "型号", "版本", "参考基价", "状态", "排序"])
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_recycle_catalog_excel_import_is_validated_and_upserts(client: TestClient):
    headers = admin_headers(client)
    template = client.get(
        "/api/admin/client/recycle/catalog/import-template", headers=headers
    )
    assert template.status_code == 200, template.text
    assert template.content.startswith(b"PK")
    assert "recycle-catalog-template.xlsx" in template.headers["content-disposition"]
    assert int(template.headers["content-length"]) == len(template.content)
    workbook = load_workbook(BytesIO(template.content), read_only=True)
    assert list(next(workbook["旧机报价"].iter_rows(values_only=True))) == [
        "品牌",
        "型号",
        "版本",
        "最高回收价",
        "状态",
        "排序",
    ]
    workbook.close()

    first = unwrap(
        client.post(
            "/api/admin/client/recycle/catalog/import",
            headers=headers,
            files={
                "file": (
                    "旧机报价.xlsx",
                    recycle_catalog_xlsx(
                        [
                            ["DJI", "Mini 4 Pro", "标准版", 2800, "启用", 10],
                            ["DJI", "Air 3S", None, "4500.50", 1, 20],
                        ]
                    ),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    )
    assert first == {
        "total": 2,
        "created": 2,
        "updated": 0,
        "items": first["items"],
    }

    second = unwrap(
        client.post(
            "/api/admin/client/recycle/catalog/import",
            headers=headers,
            files={
                "file": (
                    "price-update.xlsx",
                    recycle_catalog_xlsx(
                        [["dji", "mini 4 pro", "标准版", 2999, "停用", 30]]
                    ),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    )
    assert second["created"] == 0
    assert second["updated"] == 1
    catalog = unwrap(client.get("/api/admin/client/recycle/catalog", headers=headers))
    mini = next(item for item in catalog if item["model"].casefold() == "mini 4 pro")
    assert float(mini["reference_price"]) == 2999.0
    assert mini["enabled"] is False
    assert mini["sort_order"] == 30

    invalid = client.post(
        "/api/admin/client/recycle/catalog/import",
        headers=headers,
        files={
            "file": (
                "invalid.xlsx",
                recycle_catalog_xlsx(
                    [
                        ["DJI", "Mavic 3 Pro", None, 7000, "启用", 1],
                        ["DJI", "Avata 2", None, -1, "启用", 2],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert invalid.status_code == 422
    after_invalid = unwrap(client.get("/api/admin/client/recycle/catalog", headers=headers))
    assert all(item["model"] != "Mavic 3 Pro" for item in after_invalid)


def test_admin_can_publish_product_with_first_sku_atomically(client: TestClient):
    headers = admin_headers(client)
    category = unwrap(
        client.post(
            "/api/admin/client/product-categories",
            headers=headers,
            json={"name": "一键发布分类", "slug": "publish-flow", "enabled": True},
        ),
        201,
    )

    published = unwrap(
        client.post(
            "/api/admin/client/products/publish",
            headers=headers,
            json={
                "product": {
                    "category_id": category["id"],
                    "name": "一键发布商品",
                    "slug": "one-step-product",
                    "summary": "商品和首个规格同时创建",
                    "status": "published",
                    "featured": True,
                },
                "sku": {
                    "sku": "ONE-STEP-001",
                    "name": "标准版",
                    "attributes": {"版本": "标准版"},
                    "price": "1999.00",
                    "stock_quantity": 8,
                    "enabled": True,
                },
            },
        ),
        201,
    )

    detail = unwrap(
        client.get(f"/api/admin/client/products/{published['id']}", headers=headers)
    )
    assert detail["status"] == "published"
    assert detail["skus"][0]["id"] == published["sku_id"]
    assert float(detail["skus"][0]["price"]) == 1999.0
    public = unwrap(client.get(f"/api/client/products/{published['id']}"))
    assert public["skus"][0]["stock"] == 8


def test_publish_product_rejects_disabled_initial_sku(client: TestClient):
    headers = admin_headers(client)
    response = client.post(
        "/api/admin/client/products/publish",
        headers=headers,
        json={
            "product": {
                "name": "不可销售商品",
                "slug": "disabled-first-sku",
                "status": "published",
            },
            "sku": {
                "sku": "DISABLED-001",
                "name": "停用规格",
                "price": "10.00",
                "stock_quantity": 1,
                "enabled": False,
            },
        },
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "published_product_requires_enabled_sku"


def test_client_visible_inventory_drives_public_stock(client: TestClient):
    headers = admin_headers(client)
    category = unwrap(
        client.post(
            "/api/admin/client/product-categories",
            headers=headers,
            json={"name": "库存联动", "slug": "inventory-linked", "enabled": True},
        ),
        201,
    )
    inventory = unwrap(
        client.post(
            "/api/inventory/items",
            headers=headers,
            json={
                "sku": "PUBLIC-STOCK-001",
                "name": "可销售测试物料",
                "unit": "件",
                "purchase_price": "100.00",
                "sale_price": "199.00",
                "stock_quantity": 5,
                "safety_stock": 1,
                "enabled": True,
                "client_visible": False,
            },
        ),
        201,
    )
    hidden_publish = client.post(
        "/api/admin/client/products/publish",
        headers=headers,
        json={
            "product": {
                "category_id": category["id"],
                "name": "库存联动商品",
                "slug": "inventory-linked-product",
                "status": "published",
            },
            "sku": {
                "inventory_item_id": inventory["id"],
                "sku": "PUBLIC-STOCK-001",
                "name": "标准版",
                "price": "199.00",
                "stock_quantity": 999,
                "enabled": True,
            },
        },
    )
    assert hidden_publish.status_code == 409
    assert hidden_publish.json()["error"]["code"] == "inventory_not_client_visible"

    unwrap(
        client.patch(
            f"/api/inventory/items/{inventory['id']}/client-visibility",
            headers=headers,
            json={"client_visible": True},
        )
    )
    published = unwrap(
        client.post(
            "/api/admin/client/products/publish",
            headers=headers,
            json={
                "product": {
                    "category_id": category["id"],
                    "name": "库存联动商品",
                    "slug": "inventory-linked-product",
                    "status": "published",
                },
                "sku": {
                    "inventory_item_id": inventory["id"],
                    "sku": "PUBLIC-STOCK-001",
                    "name": "标准版",
                    "price": "199.00",
                    "stock_quantity": 999,
                    "enabled": True,
                },
            },
        ),
        201,
    )
    public = unwrap(client.get(f"/api/client/products/{published['id']}"))
    assert public["skus"][0]["stock"] == 5
    assert "purchase_price" not in public["skus"][0]
    assert "location" not in public["skus"][0]

    unwrap(
        client.post(
            "/api/inventory/transactions",
            headers=headers,
            json={
                "inventory_item_id": inventory["id"],
                "transaction_type": "stock_in",
                "quantity": 2,
                "remarks": "客户端库存联动测试",
            },
        ),
        201,
    )
    assert unwrap(client.get(f"/api/client/products/{published['id']}"))["skus"][0]["stock"] == 7

    unwrap(
        client.patch(
            f"/api/inventory/items/{inventory['id']}/client-visibility",
            headers=headers,
            json={"client_visible": False},
        )
    )
    assert client.get(f"/api/client/products/{published['id']}").status_code == 404
    listing = unwrap(client.get("/api/client/products"))
    assert all(item["id"] != published["id"] for item in listing)


def test_client_account_shop_order_and_idor(client: TestClient):
    client_page = client.get("/client")
    assert client_page.status_code == 200
    assert client_page.headers["Cache-Control"] == "no-store"
    assert client.get("/client/community").headers["Cache-Control"] == "no-store"
    asset_path = re.search(r'src="([^"]+\.js)"', client_page.text)
    assert asset_path is not None
    client_asset = client.get(asset_path.group(1))
    assert client_asset.status_code == 200
    assert client_asset.headers["Cache-Control"] == "public, max-age=31536000, immutable"
    admin_script = client.get("/static/app.js")
    assert admin_script.status_code == 200
    assert "render.client_platform" in admin_script.text
    assert "render.client_orders" in admin_script.text
    assert "render.recycle_pricing" in admin_script.text
    assert "render.forum_management" in admin_script.text
    assert "Excel 批量导入" in admin_script.text
    active_recycle_renderer = admin_script.text.rsplit(
        "render.recycle_pricing=async()=>{", 1
    )[1].split("function enhanceSearchableSelects", 1)[0]
    assert "旧机最高价报价正常使用" in active_recycle_renderer
    assert "Excel 模板下载和批量导入均正常可用" in active_recycle_renderer
    assert "/api/admin/client/recycle/catalog/import" in admin_script.text
    assert "/api/admin/client/recycle/rules" not in active_recycle_renderer
    assert "/api/admin/client/forum/comments" in admin_script.text
    staff_headers = admin_headers(client)
    category = unwrap(
        client.post(
            "/api/admin/client/product-categories",
            headers=staff_headers,
            json={"name": "测试商品", "slug": "test-products", "enabled": True},
        ),
        201,
    )
    product = unwrap(
        client.post(
            "/api/admin/client/products",
            headers=staff_headers,
            json={
                "category_id": category["id"],
                "name": "DJI 测试配件",
                "slug": "dji-test-accessory",
                "summary": "集成测试商品",
                "status": "published",
                "featured": True,
            },
        ),
        201,
    )
    sku = unwrap(
        client.post(
            f"/api/admin/client/products/{product['id']}/skus",
            headers=staff_headers,
            json={
                "sku": "CLIENT-TEST-001",
                "name": "标准版",
                "attributes": {"version": "standard"},
                "price": "1299.00",
                "stock_quantity": 3,
                "enabled": True,
            },
        ),
        201,
    )
    detail = unwrap(client.get(f"/api/admin/client/products/{product['id']}", headers=staff_headers))
    assert detail["skus"][0]["id"] == sku["id"]
    updated_sku = unwrap(
        client.patch(
            f"/api/admin/client/skus/{sku['id']}",
            headers=staff_headers,
            json={
                "sku": "CLIENT-TEST-001",
                "name": "标准版",
                "attributes": {"version": "standard"},
                "price": "1299.00",
                "stock_quantity": 4,
                "enabled": True,
            },
        )
    )
    assert updated_sku["stock_quantity"] == 4
    product_image = unwrap(
        client.post(
            f"/api/admin/client/products/{product['id']}/images",
            headers=staff_headers,
            files={"file": ("product.png", png_bytes(), "image/png")},
        ),
        201,
    )
    assert client.get(product_image["url"]).status_code == 200
    unwrap(
        client.delete(
            f"/api/admin/client/product-images/{product_image['id']}",
            headers=staff_headers,
        )
    )
    assert client.get(product_image["url"]).status_code == 404
    updated_product = unwrap(
        client.patch(
            f"/api/admin/client/products/{product['id']}",
            headers=staff_headers,
            json={
                "category_id": category["id"],
                "name": "DJI 测试配件",
                "slug": "dji-test-accessory",
                "summary": "已更新的集成测试商品",
                "status": "published",
                "featured": True,
            },
        )
    )
    assert updated_product["summary"] == "已更新的集成测试商品"

    account_a, cookie_a = register(client, "1")
    assert account_a["identifier"] == "@client_1"
    csrf_a = account_a["csrf_token"]
    address = unwrap(
        client.post(
            "/api/client/addresses",
            headers={"X-CSRF-Token": csrf_a},
            json={
                "recipient_name": "客户一",
                "phone": "13800000001",
                "province": "广东省",
                "city": "深圳市",
                "district": "南山区",
                "detail": "测试路 1 号",
                "is_default": True,
            },
        ),
        201,
    )
    public_product = unwrap(client.get(f"/api/client/products/{product['id']}"))
    assert public_product["skus"][0]["price"] == "1299.00"
    assert "inventory_item_id" not in public_product["skus"][0]
    cart = unwrap(
        client.post(
            "/api/client/cart/items",
            headers={"X-CSRF-Token": csrf_a},
            json={"sku_id": sku["id"], "quantity": 2},
        ),
        201,
    )
    order = unwrap(
        client.post(
            "/api/client/orders",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "order-client-a-001"},
            json={
                "address_id": address["id"],
                "delivery_method": "shipping",
                "cart_item_ids": [cart["items"][0]["id"]],
            },
        ),
        201,
    )
    assert order["total_amount"] == "2598.00"
    assert order["payment_provider"] == "manual"
    duplicate_order = unwrap(
        client.post(
            "/api/client/orders",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "order-client-a-001"},
            json={
                "address_id": address["id"],
                "delivery_method": "shipping",
                "cart_item_ids": [cart["items"][0]["id"]],
            },
        ),
        201,
    )
    assert duplicate_order["id"] == order["id"]
    work_items = unwrap(client.get("/api/client/work-items"))
    assert [(item["type"], item["id"]) for item in work_items] == [("retail", order["id"])]

    account_b, _cookie_b = register(client, "2")
    forbidden = client.get(f"/api/client/orders/{order['id']}")
    assert forbidden.status_code == 404

    client.cookies.set("service_client_session", cookie_a)
    client.headers["X-CSRF-Token"] = csrf_a
    assert unwrap(client.get(f"/api/client/orders/{order['id']}"))["id"] == order["id"]
    assert unwrap(client.get("/api/client/me"))["account"]["identifier"] == "@client_1"


def test_repair_recycle_forum_upload_and_security(client: TestClient):
    staff_headers = admin_headers(client)
    forum_category = unwrap(
        client.post(
            "/api/admin/client/forum/categories",
            headers=staff_headers,
            json={"name": "维修交流", "slug": "repair-test", "enabled": True},
        ),
        201,
    )
    catalog = unwrap(
        client.post(
            "/api/admin/client/recycle/catalog",
            headers=staff_headers,
            json={
                "brand": "DJI",
                "model": "Air 3S",
                "variant": "标准版",
                "reference_price": "5000.00",
                "enabled": True,
            },
        ),
        201,
    )
    disabled_rule_payload = {
        "code": "appearance_good_test",
        "rule_group": "appearance",
        "label": "正常使用痕迹",
        "factor": "0.8000",
        "adjustment": "0.00",
        "enabled": True,
    }
    disabled_rules = client.get(
        "/api/admin/client/recycle/rules", headers=staff_headers
    )
    assert disabled_rules.status_code == 410
    assert disabled_rules.json()["error"] == {
        "code": "recycle_rules_disabled",
        "message": "仅旧报价规则已停用；旧机最高价维护和 Excel 批量导入正常可用",
    }
    assert client.post(
        "/api/admin/client/recycle/rules",
        headers=staff_headers,
        json=disabled_rule_payload,
    ).status_code == 410
    assert client.patch(
        "/api/admin/client/recycle/rules/1",
        headers=staff_headers,
        json=disabled_rule_payload,
    ).status_code == 410

    account_a, cookie_a = register(client, "11")
    csrf_a = account_a["csrf_token"]
    avatar = unwrap(
        client.post(
            "/api/client/me/avatar",
            headers={"X-CSRF-Token": csrf_a},
            files={"file": ("avatar.png", png_bytes(), "image/png")},
        )
    )
    assert avatar["avatar_url"].startswith(f"/api/client/avatars/{account_a['id']}?v=")
    assert client.get(avatar["avatar_url"]).content == png_bytes()
    address = unwrap(
        client.post(
            "/api/client/addresses",
            headers={"X-CSRF-Token": csrf_a},
            json={
                "recipient_name": "维修客户",
                "phone": "13800000011",
                "province": "广东省",
                "city": "深圳市",
                "detail": "维修路 11 号",
                "is_default": True,
            },
        ),
        201,
    )
    repair_payload = {
        "brand": "DJI",
        "model": "Air 3S",
        "serial_number": "CLIENT-SN-001",
        "fault_type": "云台异常",
        "fault_description": "云台启动后持续抖动，无法稳定画面",
        "service_mode": "shipping",
        "has_crash_damage": True,
        "accessories": ["电池", "遥控器"],
        "contact_name": "维修客户",
        "contact_phone": "13800000011",
        "address_id": address["id"],
    }
    repair = unwrap(
        client.post(
            "/api/client/repair",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "repair-client-a-001"},
            json=repair_payload,
        ),
        201,
    )
    assert repair["status_label"] == "申请已提交"
    duplicate_repair = unwrap(
        client.post(
            "/api/client/repair",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "repair-client-a-001"},
            json=repair_payload,
        ),
        201,
    )
    assert duplicate_repair["id"] == repair["id"]
    quote = unwrap(
        client.post(
            "/api/quotes",
            headers=staff_headers,
            json={
                "repair_order_id": repair["id"],
                "labor_fee": "100.00",
                "shipping_fee": "20.00",
                "customer_notice": "测试报价仅用于自动化验收",
                "items": [
                    {
                        "item_name": "云台检测与维修",
                        "quantity": "1",
                        "unit_price": "680.00",
                        "cost_price": "0.00",
                        "item_type": "service",
                    }
                ],
            },
        ),
        201,
    )
    repair_with_quote = unwrap(client.get(f"/api/client/repair/{repair['id']}"))
    assert repair_with_quote["current_quote"]["id"] == quote["id"]
    quote_pdf = client.get(f"/api/client/repair/quotes/{quote['id']}/pdf")
    assert quote_pdf.status_code == 200, quote_pdf.text
    assert quote_pdf.headers["content-type"].startswith("application/pdf")
    assert quote_pdf.content.startswith(b"%PDF")
    accepted = unwrap(
        client.post(
            f"/api/client/repair/{repair['id']}/quote-decision",
            headers={"X-CSRF-Token": csrf_a},
            json={"decision": "accepted"},
        )
    )
    assert accepted["status"] == "customer_confirmed"
    assert accepted["current_quote"]["status"] == "confirmed"
    image = png_bytes()
    upload = unwrap(
        client.post(
            f"/api/client/repair/{repair['id']}/attachments",
            headers={"X-CSRF-Token": csrf_a},
            files={"file": ("damage.png", image, "image/png")},
        ),
        201,
    )
    assert client.get(upload["url"]).content == image
    invalid = client.post(
        f"/api/client/repair/{repair['id']}/attachments",
        headers={"X-CSRF-Token": csrf_a},
        files={"file": ("fake.png", b"not-an-image", "image/png")},
    )
    assert invalid.status_code == 415

    disabled_client_rules = client.get("/api/client/recycle/rules")
    assert disabled_client_rules.status_code == 410
    assert disabled_client_rules.json()["error"] == {
        "code": "recycle_rules_disabled",
        "message": "仅旧报价规则已停用；旧机最高价报价功能正常可用",
    }
    old_rule_estimate = client.post(
        "/api/client/recycle/estimate",
        json={
            "catalog_item_id": catalog["id"],
            "condition_codes": ["appearance_good_test"],
            "details": {},
        },
    )
    assert old_rule_estimate.status_code == 410
    estimate = unwrap(
        client.post(
            "/api/client/recycle/estimate",
            json={
                "catalog_item_id": catalog["id"],
                "condition_codes": [],
                "details": {},
            },
        )
    )
    assert estimate["maximum_price"] == "5000.00"
    assert estimate["reference_min"] == "5000.00"
    assert estimate["reference_max"] == "5000.00"
    recycle = unwrap(
        client.post(
            "/api/client/recycle",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "recycle-client-a-001"},
            json={
                "catalog_item_id": catalog["id"],
                "condition_codes": [],
                "details": {},
                "contact_name": "回收客户",
                "contact_phone": "13800000011",
                "contact_wechat": "recycle-client-11",
                "device_condition": "设备可正常开机，机身有轻微使用痕迹",
                "notes": "工作日下午方便联系",
                "submit": True,
            },
        ),
        201,
    )
    assert recycle["status"] == "submitted"
    assert recycle["maximum_price"] == "5000.00"
    assert recycle["contact_phone"] == "13800000011"
    backend_requests = unwrap(
        client.get("/api/admin/client/recycle/requests", headers=staff_headers)
    )
    backend_recycle = next(item for item in backend_requests if item["id"] == recycle["id"])
    assert backend_recycle["catalog_item"]["model"] == "Air 3S"
    assert backend_recycle["contact_name"] == "回收客户"
    assert backend_recycle["contact_phone"] == "13800000011"
    assert backend_recycle["device_condition"] == "设备可正常开机，机身有轻微使用痕迹"

    post = unwrap(
        client.post(
            "/api/client/forum/posts",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "forum-post-client-a-001"},
            json={
                "category_id": forum_category["id"],
                "title": "云台抖动排查交流",
                "content": "设备启动后云台抖动，已提交维修检查。",
            },
        ),
        201,
    )
    assert post["author"]["avatar_url"] == avatar["avatar_url"]
    assert post["author"]["username"] == account_a["username"]
    assert post["author"]["identifier"] == "@client_11"
    public_posts = unwrap(client.get("/api/client/forum/posts"))["items"]
    assert public_posts[0]["author"]["avatar_url"] == avatar["avatar_url"]
    comment = unwrap(
        client.post(
            f"/api/client/forum/posts/{post['id']}/comments",
            headers={"X-CSRF-Token": csrf_a, "Idempotency-Key": "forum-comment-a-001"},
            json={"content": "补充：设备有轻微碰撞。"},
        ),
        201,
    )
    assert unwrap(client.get(f"/api/client/forum/posts/{post['id']}"))["post"]["comment_count"] == 1
    merged = unwrap(client.get("/api/client/work-items"))
    assert {item["type"] for item in merged} == {"repair", "recycle"}
    comments = unwrap(client.get("/api/admin/client/forum/comments", headers=staff_headers))
    assert any(item["id"] == comment["id"] for item in comments)
    unwrap(
        client.patch(
            f"/api/admin/client/forum/comments/{comment['id']}",
            headers=staff_headers,
            json={"status": "hidden"},
        )
    )
    assert unwrap(client.get(f"/api/client/forum/posts/{post['id']}"))["post"]["comment_count"] == 0
    unwrap(
        client.patch(
            f"/api/admin/client/forum/comments/{comment['id']}",
            headers=staff_headers,
            json={"status": "published"},
        )
    )

    account_b, _cookie_b = register(client, "12")
    csrf_b = account_b["csrf_token"]
    forbidden_edit = client.patch(
        f"/api/client/forum/posts/{post['id']}",
        headers={"X-CSRF-Token": csrf_b},
        json={"title": "越权修改"},
    )
    assert forbidden_edit.status_code == 403
    repair_idor = client.get(f"/api/client/repair/{repair['id']}")
    assert repair_idor.status_code == 404

    client.cookies.set("service_client_session", "tampered-session")
    invalid_session = client.get("/api/client/me")
    assert invalid_session.status_code == 401
    client.cookies.set("service_client_session", cookie_a)


def test_client_replacement_creates_backend_ticket_and_is_idempotent(client: TestClient):
    staff_headers = admin_headers(client)
    account_a, cookie_a = register(client, "61")
    csrf_a = account_a["csrf_token"]
    address = unwrap(
        client.post(
            "/api/client/addresses",
            headers={"X-CSRF-Token": csrf_a},
            json={
                "recipient_name": "置换客户",
                "phone": "13800000061",
                "province": "广东省",
                "city": "深圳市",
                "district": "南山区",
                "detail": "服务路 61 号",
                "is_default": True,
            },
        ),
        201,
    )
    payload = {
        "old_model": "DJI Mini 4 Pro 畅飞套装",
        "desired_model": "DJI Air 3S 畅飞套装",
        "contact_name": "置换客户",
        "contact_phone": "13800000061",
        "address_id": address["id"],
        "notes": "工作日下午方便联系",
    }
    replacement = unwrap(
        client.post(
            "/api/client/replacement",
            headers={
                "X-CSRF-Token": csrf_a,
                "Idempotency-Key": "replacement-client-a-001",
            },
            json=payload,
        ),
        201,
    )
    assert replacement["status"] == "open"
    assert replacement["status_label"] == "已提交"
    assert replacement["old_model"] == payload["old_model"]
    assert replacement["desired_model"] == payload["desired_model"]
    assert replacement["address"]["detail"] == "服务路 61 号"
    assert replacement["notice"] == "服务顾问会在一个工作日内联系您，请保持电话畅通"

    duplicate = unwrap(
        client.post(
            "/api/client/replacement",
            headers={
                "X-CSRF-Token": csrf_a,
                "Idempotency-Key": "replacement-client-a-001",
            },
            json=payload,
        ),
        201,
    )
    assert duplicate["id"] == replacement["id"]
    assert unwrap(client.get(f"/api/client/replacement/{replacement['id']}"))["notes"] == payload["notes"]
    work_items = unwrap(client.get("/api/client/work-items"))
    assert any(
        item["type"] == "replacement" and item["id"] == replacement["id"]
        for item in work_items
    )
    assert unwrap(client.get("/api/client/me"))["counts"]["replacements"] == 1

    backend = unwrap(
        client.get("/api/service-tickets?ticket_type=replacement", headers=staff_headers)
    )
    backend_ticket = next(item for item in backend if item["id"] == replacement["id"])
    assert backend_ticket["customer_id"] == account_a["customer_id"]
    assert "旧机型：DJI Mini 4 Pro 畅飞套装" in backend_ticket["description"]
    assert "联系地址：广东省深圳市南山区服务路 61 号" in backend_ticket["description"]

    register(client, "62")
    assert client.get(f"/api/client/replacement/{replacement['id']}").status_code == 404
    client.cookies.set("service_client_session", cookie_a)
    client.headers["X-CSRF-Token"] = csrf_a


def test_client_identifier_can_change_twice_per_calendar_year(client: TestClient):
    account_a, cookie_a = register(client, "71")
    csrf_a = account_a["csrf_token"]
    initial = unwrap(client.get("/api/client/me"))["identifier_change"]
    assert initial["limit"] == 2
    assert initial["used"] == 0
    assert initial["remaining"] == 2

    unchanged = client.patch(
        "/api/client/auth/identifier",
        headers={"X-CSRF-Token": csrf_a},
        json={"identifier": "@client_71"},
    )
    assert unchanged.status_code == 409
    assert unchanged.json()["error"]["code"] == "identifier_unchanged"

    first = unwrap(
        client.patch(
            "/api/client/auth/identifier",
            headers={"X-CSRF-Token": csrf_a},
            json={"identifier": "@First.Identifier"},
        )
    )
    assert first["account"]["username"] == "first.identifier"
    assert first["account"]["identifier"] == "@first.identifier"
    assert first["identifier_change"]["remaining"] == 1

    account_b, _cookie_b = register(client, "72")
    conflict = client.patch(
        "/api/client/auth/identifier",
        headers={"X-CSRF-Token": account_b["csrf_token"]},
        json={"identifier": "FIRST.IDENTIFIER"},
    )
    assert conflict.status_code == 409
    assert conflict.json()["error"]["code"] == "identifier_conflict"
    assert unwrap(client.get("/api/client/me"))["identifier_change"]["remaining"] == 2

    client.cookies.set("service_client_session", cookie_a)
    client.headers["X-CSRF-Token"] = csrf_a
    second = unwrap(
        client.patch(
            "/api/client/auth/identifier",
            json={"identifier": "second_identifier"},
        )
    )
    assert second["account"]["identifier"] == "@second_identifier"
    assert second["identifier_change"]["used"] == 2
    assert second["identifier_change"]["remaining"] == 0

    exhausted = client.patch(
        "/api/client/auth/identifier", json={"identifier": "third_identifier"}
    )
    assert exhausted.status_code == 429
    assert exhausted.json()["error"]["code"] == "identifier_change_limit_reached"
    assert unwrap(client.get("/api/client/me"))["identifier_change"]["remaining"] == 0

    unwrap(client.post("/api/client/auth/logout"))
    assert client.post(
        "/api/client/auth/login",
        json={"login": "client_71", "password": "ClientPass123"},
    ).status_code == 401
    relogin = unwrap(
        client.post(
            "/api/client/auth/login",
            json={"login": "@second_identifier", "password": "ClientPass123"},
        )
    )
    assert relogin["identifier"] == "@second_identifier"


def test_forum_feed_personalizes_tracks_dwell_and_respects_negative_feedback(
    client: TestClient,
):
    staff_headers = admin_headers(client)
    category = unwrap(
        client.post(
            "/api/admin/client/forum/categories",
            headers=staff_headers,
            json={"name": "推荐算法测试", "slug": "ranking-test", "enabled": True},
        ),
        201,
    )
    author, _author_cookie = register(client, "31")
    first = unwrap(
        client.post(
            "/api/client/forum/posts",
            headers={
                "X-CSRF-Token": author["csrf_token"],
                "Idempotency-Key": "recommendation-post-001",
            },
            json={
                "category_id": category["id"],
                "title": "云台排查经验一",
                "content": "记录云台抖动的完整排查过程。",
            },
        ),
        201,
    )
    second = unwrap(
        client.post(
            "/api/client/forum/posts",
            headers={
                "X-CSRF-Token": author["csrf_token"],
                "Idempotency-Key": "recommendation-post-002",
            },
            json={
                "category_id": category["id"],
                "title": "云台排查经验二",
                "content": "补充另一种排查思路。",
            },
        ),
        201,
    )
    viewer, _viewer_cookie = register(client, "32")

    initial = unwrap(client.get("/api/client/forum/posts?sort=recommended"))
    assert initial["feed"]["strategy"] == "multi_signal_v1"
    assert initial["feed"]["personalized"] is False
    assert {first["id"], second["id"]}.issubset(
        {item["id"] for item in initial["items"]}
    )
    assert all(item["recommendation_reason"] for item in initial["items"])

    recorded = unwrap(
        client.post(
            "/api/client/forum/signals",
            headers={"X-CSRF-Token": viewer["csrf_token"]},
            json={
                "items": [
                    {
                        "post_id": first["id"],
                        "impression": True,
                        "dwell_time_ms": 4200,
                    }
                ]
            },
        )
    )
    assert recorded == {"recorded": 1}
    personalized = unwrap(client.get("/api/client/forum/posts?sort=recommended"))
    assert personalized["feed"]["personalized"] is True
    assert "阅读" in personalized["feed"]["description"]

    unwrap(
        client.post(
            "/api/client/forum/signals",
            headers={"X-CSRF-Token": viewer["csrf_token"]},
            json={"items": [{"post_id": first["id"], "not_interested": True}]},
        )
    )
    assert first["id"] not in {
        item["id"]
        for item in unwrap(client.get("/api/client/forum/posts?sort=recommended"))["items"]
    }
    assert first["id"] not in {
        item["id"]
        for item in unwrap(client.get("/api/client/forum/posts?sort=latest"))["items"]
    }

    unwrap(
        client.post(
            "/api/client/forum/signals",
            headers={"X-CSRF-Token": viewer["csrf_token"]},
            json={"items": [{"post_id": first["id"], "not_interested": False}]},
        )
    )
    assert first["id"] in {
        item["id"]
        for item in unwrap(client.get("/api/client/forum/posts?sort=recommended"))["items"]
    }

    duplicate = client.post(
        "/api/client/forum/signals",
        headers={"X-CSRF-Token": viewer["csrf_token"]},
        json={
            "items": [
                {"post_id": first["id"], "impression": True},
                {"post_id": first["id"], "dwell_time_ms": 2000},
            ]
        },
    )
    assert duplicate.status_code == 422
