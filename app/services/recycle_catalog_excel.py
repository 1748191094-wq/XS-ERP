from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import ValidationError

from app.core.exceptions import BusinessError
from app.schemas.client import RecycleCatalogWrite


MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 2000


@dataclass(frozen=True)
class RecycleCatalogImportRow:
    row_number: int
    payload: RecycleCatalogWrite

    @property
    def key(self) -> tuple[str, str, str]:
        return (
            self.payload.brand.casefold(),
            self.payload.model.casefold(),
            (self.payload.variant or "").casefold(),
        )


HEADER_ALIASES = {
    "brand": {"品牌", "brand"},
    "model": {"型号", "model"},
    "variant": {"版本", "容量版本", "variant"},
    "reference_price": {
        "最高价", "最高回收价", "参考价", "参考基价", "回收参考价", "回收参考基价",
        "price", "referenceprice", "maximumprice",
    },
    "enabled": {"状态", "是否启用", "启用", "enabled"},
    "sort_order": {"排序", "排序值", "sort", "sortorder"},
}


def _normal_header(value: object) -> str:
    return re.sub(r"[\s_\-/]+", "", str(value or "")).casefold()


def _header_map(values: tuple[object, ...]) -> dict[str, int]:
    alias_lookup = {
        _normal_header(alias): field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        field = alias_lookup.get(_normal_header(value))
        if field and field not in result:
            result[field] = index
    missing = [label for field, label in (("model", "型号"), ("reference_price", "最高回收价")) if field not in result]
    if missing:
        raise BusinessError(
            f"Excel 缺少必填列：{'、'.join(missing)}",
            code="recycle_catalog_import_headers_invalid",
            status_code=422,
        )
    return result


def _enabled(value: object) -> bool:
    if value is None or str(value).strip() == "":
        return True
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().casefold()
    if normalized in {"1", "true", "yes", "y", "是", "启用", "上架"}:
        return True
    if normalized in {"0", "false", "no", "n", "否", "停用", "下架"}:
        return False
    raise ValueError("状态只能填写启用/停用、是/否或 1/0")


def _cell(values: tuple[object, ...], headers: dict[str, int], field: str) -> object | None:
    index = headers.get(field)
    return values[index] if index is not None and index < len(values) else None


def parse_recycle_catalog_workbook(content: bytes) -> list[RecycleCatalogImportRow]:
    if not content:
        raise BusinessError("Excel 文件为空", code="recycle_catalog_import_empty", status_code=422)
    if len(content) > MAX_IMPORT_BYTES:
        raise BusinessError("Excel 文件不能超过 5 MB", code="recycle_catalog_import_too_large", status_code=413)
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > 200 or sum(member.file_size for member in members) > 50 * 1024 * 1024:
                raise BusinessError(
                    "Excel 解压后内容过大",
                    code="recycle_catalog_import_expanded_too_large",
                    status_code=413,
                )
    except BadZipFile as exc:
        raise BusinessError(
            "无法读取 Excel，请使用系统模板并保存为 .xlsx 文件",
            code="recycle_catalog_import_invalid_file",
            status_code=422,
        ) from exc
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise BusinessError(
            "无法读取 Excel，请使用系统模板并保存为 .xlsx 文件",
            code="recycle_catalog_import_invalid_file",
            status_code=422,
        ) from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first_row = next(iterator, None)
        if first_row is None:
            raise BusinessError("Excel 没有表头", code="recycle_catalog_import_empty", status_code=422)
        headers = _header_map(first_row)
        parsed: list[RecycleCatalogImportRow] = []
        errors: list[str] = []
        seen: dict[tuple[str, str, str], int] = {}
        for row_number, values in enumerate(iterator, start=2):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            if len(parsed) >= MAX_IMPORT_ROWS:
                raise BusinessError(
                    f"单次最多导入 {MAX_IMPORT_ROWS} 行",
                    code="recycle_catalog_import_too_many_rows",
                    status_code=422,
                )
            try:
                brand = str(_cell(values, headers, "brand") or "DJI").strip()
                model = str(_cell(values, headers, "model") or "").strip()
                variant_value = _cell(values, headers, "variant")
                variant = str(variant_value).strip() if variant_value is not None and str(variant_value).strip() else None
                price_value = _cell(values, headers, "reference_price")
                try:
                    price = Decimal(str(price_value).strip())
                except (InvalidOperation, ValueError, AttributeError):
                    raise ValueError("最高回收价必须是有效数字") from None
                sort_value = _cell(values, headers, "sort_order")
                if sort_value is None or str(sort_value).strip() == "":
                    sort_order = 0
                else:
                    try:
                        numeric_sort = Decimal(str(sort_value).strip())
                    except (InvalidOperation, ValueError, AttributeError):
                        raise ValueError("排序必须是整数") from None
                    if not numeric_sort.is_finite():
                        raise ValueError("排序必须是整数")
                    if numeric_sort != numeric_sort.to_integral_value():
                        raise ValueError("排序必须是整数")
                    sort_order = int(numeric_sort)
                payload = RecycleCatalogWrite(
                    brand=brand,
                    model=model,
                    variant=variant,
                    reference_price=price,
                    enabled=_enabled(_cell(values, headers, "enabled")),
                    sort_order=sort_order,
                )
                item = RecycleCatalogImportRow(row_number=row_number, payload=payload)
                if item.key in seen:
                    raise ValueError(f"与第 {seen[item.key]} 行的品牌、型号和版本重复")
                seen[item.key] = row_number
                parsed.append(item)
            except (ValueError, ValidationError) as exc:
                if isinstance(exc, ValidationError):
                    issue = exc.errors()[0].get("msg", "字段格式不正确")
                else:
                    issue = str(exc)
                errors.append(f"第 {row_number} 行：{issue}")
        if errors:
            preview = "；".join(errors[:12])
            suffix = f"；另有 {len(errors) - 12} 行错误" if len(errors) > 12 else ""
            raise BusinessError(
                f"Excel 校验失败：{preview}{suffix}",
                code="recycle_catalog_import_rows_invalid",
                status_code=422,
            )
        if not parsed:
            raise BusinessError("Excel 没有可导入的数据行", code="recycle_catalog_import_no_rows", status_code=422)
        return parsed
    finally:
        workbook.close()


def build_recycle_catalog_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "旧机报价"
    headers = ["品牌", "型号", "版本", "最高回收价", "状态", "排序"]
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = [14, 24, 20, 16, 14, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    notes = workbook.create_sheet("填写说明")
    notes.append(["字段", "要求"])
    notes.append(["品牌", "选填，留空默认 DJI"])
    notes.append(["型号", "必填，例如 Mini 4 Pro"])
    notes.append(["版本", "选填，例如 畅飞套装"])
    notes.append(["最高回收价", "必填，必须大于或等于 0；客户端仅展示此价格"])
    notes.append(["状态", "选填：启用/停用、是/否或 1/0；留空默认启用"])
    notes.append(["排序", "选填，0 至 10000 的整数；留空默认 0"])
    notes.append(["更新规则", "品牌、型号、版本相同则更新现有报价，否则新增"])
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 68
    notes.freeze_panes = "A2"
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()
